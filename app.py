import os
import uuid
from datetime import date, datetime
from decimal import Decimal
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_from_directory
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
load_dotenv()
load_dotenv(BASE_DIR / "finan.env")

UPLOAD_DIR = BASE_DIR / os.getenv("UPLOAD_DIR", "uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

DATABASE_URL = os.getenv("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

if not DATABASE_URL:
    DATABASE_URL = f"sqlite:///{BASE_DIR / 'app.db'}"

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret")

CORS(app)
db = SQLAlchemy(app)

TOKENS = {}

ROLES = {"cashier", "board", "auditor", "admin", "member"}
EXPENSE_CATEGORIES = [
    "Saimnieciskie izdevumi",
    "Būvmateriāli",
    "Piebarošana",
    "Nodokļi",
    "Licences",
    "Platību maksājumi",
    "Internets",
    "Elektrība",
    "Apdrošināšana",
    "Pļaušanas izdevumi",
    "Bebru uzraudzības izdevumi",
    "LMS biedru maksa",
    "Bankas komisijas maksa",
    "Citi",
]


class Member(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    list_no = db.Column(db.Integer, nullable=False)
    first_name = db.Column(db.String(80), nullable=False)
    last_name = db.Column(db.String(80), nullable=False)
    phone = db.Column(db.String(8), unique=True, nullable=False)
    status = db.Column(db.String(50), nullable=False, default="active")
    membership_fee = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    paid_this_period = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    joining_fee_paid = db.Column("Iestāšanās maksa", db.Boolean, nullable=False, default=False)
    role = db.Column(db.String(20), nullable=False, default="member")
    pin_hash = db.Column(db.String(255), nullable=True)


class MemberStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)


class MemberSeasonFee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey("member.id"), nullable=False)
    season_label = db.Column(db.String(9), nullable=False)
    membership_fee = db.Column(db.Numeric(10, 2), nullable=False, default=0)

    __table_args__ = (db.UniqueConstraint("member_id", "season_label", name="uq_member_season_fee"),)


class Period(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    season_label = db.Column(db.String(9), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    carry_over = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    active = db.Column(db.Boolean, nullable=False, default=True)


class PeriodLock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    season_label = db.Column(db.String(9), unique=True, nullable=False)
    membership_fee_locked = db.Column(db.Boolean, nullable=False, default=False)
    carry_over_locked = db.Column(db.Boolean, nullable=False, default=False)


class Income(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    income_type = db.Column(db.String(20), nullable=False)
    member_id = db.Column(db.Integer, db.ForeignKey("member.id"), nullable=True)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    entry_date = db.Column(db.Date, nullable=False)
    description = db.Column(db.String(255), nullable=True)


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(50), nullable=False)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    entry_date = db.Column(db.Date, nullable=False)
    description = db.Column(db.String(255), nullable=True)
    attachment = db.Column(db.String(255), nullable=True)
    created_by_member_id = db.Column(db.Integer, db.ForeignKey("member.id"), nullable=True)


def to_decimal(value):
    return Decimal(str(value)).quantize(Decimal("0.01"))


def current_period():
    period = Period.query.filter_by(active=True).first()
    if period:
        return period

    today = date.today()
    year = today.year if today.month >= 4 else today.year - 1
    period = Period(
        season_label=f"{year}/{year + 1}",
        start_date=date(year, 4, 1),
        end_date=date(year + 1, 3, 31),
        carry_over=0,
        active=True,
    )
    db.session.add(period)
    db.session.commit()
    return period


def get_or_create_period_lock(season_label):
    period_lock = PeriodLock.query.filter_by(season_label=season_label).first()
    if period_lock:
        return period_lock

    period_lock = PeriodLock(season_label=season_label)
    db.session.add(period_lock)
    db.session.flush()
    return period_lock


def get_period_for_request():
    season_label = (request.args.get("season_label") or "").strip()
    if season_label:
        period = Period.query.filter_by(season_label=season_label).order_by(Period.id.desc()).first()
        if period:
            return period
    return current_period()


def period_totals(period):
    income_sum = db.session.query(db.func.coalesce(db.func.sum(Income.amount), 0)).filter(
        Income.entry_date >= period.start_date,
        Income.entry_date <= period.end_date,
    ).scalar()
    expense_sum = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)).filter(
        Expense.entry_date >= period.start_date,
        Expense.entry_date <= period.end_date,
    ).scalar()

    income_total = to_decimal(income_sum) + to_decimal(period.carry_over)
    expense_total = to_decimal(expense_sum)
    diff = income_total - expense_total

    return {
        "income_total": float(income_total),
        "expense_total": float(expense_total),
        "difference": float(diff),
        "balance": float(diff) if diff >= 0 else 0.0,
        "deficit": float(diff) if diff < 0 else 0.0,
    }
def member_to_dict(member, paid_this_period_override=None, membership_fee_override=None):
    normalized_role = normalize_role(member.role)
    paid_this_period = (
        float(paid_this_period_override)
        if paid_this_period_override is not None
        else float(member.paid_this_period)
    )
    membership_fee = (
        float(membership_fee_override)
        if membership_fee_override is not None
        else float(member.membership_fee)
    )
    return {
        "id": member.id,
        "list_no": member.list_no,
        "first_name": member.first_name,
        "last_name": member.last_name,
        "phone": member.phone,
        "status": member.status,
        "membership_fee": membership_fee,
        "paid_this_period": paid_this_period,
        "joining_fee_paid": bool(member.joining_fee_paid),
        "role": normalized_role,
        "has_pin": member.pin_hash is not None,
    }


def normalize_role(raw_role):
    role = str(raw_role or "").strip().lower()
    aliases = {
        "kasieris": "cashier",
        "cashier": "cashier",
        "valde": "board",
        "board": "board",
        "revizors": "auditor",
        "auditors": "auditor",
        "auditor": "auditor",
        "admins": "admin",
        "admin": "admin",
        "biedrs": "member",
        "member": "member",
    }
    return aliases.get(role, role)


def is_admin_member(member):
    return normalize_role(member.role) == "admin"


def get_member_season_fee(member_id, season_label):
    return MemberSeasonFee.query.filter_by(member_id=member_id, season_label=season_label).first()


def get_effective_membership_fee(member, season_label):
    if season_label:
        fee_row = get_member_season_fee(member.id, season_label)
        if fee_row:
            return to_decimal(fee_row.membership_fee)
    return to_decimal(member.membership_fee)


def set_membership_fee_for_season(member, season_label, membership_fee):
    fee_value = to_decimal(membership_fee)
    fee_row = get_member_season_fee(member.id, season_label)
    if fee_row:
        fee_row.membership_fee = fee_value
    else:
        db.session.add(
            MemberSeasonFee(
                member_id=member.id,
                season_label=season_label,
                membership_fee=fee_value,
            )
        )
    return fee_value


def token_required(allowed_roles=None):
    allowed_roles = allowed_roles or ROLES

    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            auth = request.headers.get("Authorization", "")
            token = auth.replace("Bearer ", "").strip()
            user_id = TOKENS.get(token)
            if not user_id:
                return jsonify({"error": "Nepieciešama autorizācija"}), 401

            user = db.session.get(Member, user_id)
            if not user:
                return jsonify({"error": "Lietotājs nav atrasts"}), 401

            user_role = normalize_role(user.role)
            if user_role not in allowed_roles:
                return jsonify({"error": "Nepietiekamas tiesības"}), 403

            request.current_user = user
            return fn(*args, **kwargs)

        return wrapper

    return decorator


def validate_phone(phone):
    return phone.isdigit() and len(phone) == 8


def validate_pin(pin):
    return pin.isdigit() and len(pin) == 4


def to_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on", "ja"}
    return bool(value)


def calculate_membership_fee_for_period(member, base_fee):
    status = (member.status or "").strip().lower()
    base = to_decimal(base_fee)
    fee = base

    if status == "vip":
        fee = Decimal("0.00")
    elif status == "vecbiedrs 2/3":
        fee = (base * Decimal("2") / Decimal("3")).quantize(Decimal("0.01"))
    elif status == "vecbiedrs 1/2":
        fee = (base / Decimal("2")).quantize(Decimal("0.01"))

    # Ja iestāšanās maksa ir atzīmēta, periodā pieskaita vēl 2 pilnas gada biedra maksas.
    if to_bool(member.joining_fee_paid):
        fee += (base * Decimal("2")).quantize(Decimal("0.01"))

    return fee.quantize(Decimal("0.01"))


def resequence_members():
    members = Member.query.order_by(Member.list_no, Member.id).all()
    for index, member in enumerate(members, start=1):
        member.list_no = index


def ensure_unique_member_numbers():
    members = Member.query.order_by(Member.list_no, Member.id).all()
    changed = False
    for index, member in enumerate(members, start=1):
        if member.list_no != index:
            member.list_no = index
            changed = True
    return changed


def next_member_list_no():
    current_max = db.session.query(db.func.coalesce(db.func.max(Member.list_no), 0)).scalar()
    return int(current_max) + 1


def ensure_seed_data():
    default_statuses = ["Biedrs", "Kandidāts", "VIP", "Vecbiedrs 2/3", "Vecbiedrs 1/2"]
    for name in default_statuses:
        if not MemberStatus.query.filter_by(name=name).first():
            db.session.add(MemberStatus(name=name))
    db.session.commit()

    if Member.query.count() == 0:
        admin = Member(
            list_no=1,
            first_name="Admin",
            last_name="Konts",
            phone="29123456",
            status="active",
            membership_fee=0,
            paid_this_period=0,
            role="admin",
            pin_hash=generate_password_hash("0308"),
        )
        db.session.add(admin)
        db.session.commit()

    period = current_period()

    for member in Member.query.all():
        if not get_member_season_fee(member.id, period.season_label):
            db.session.add(
                MemberSeasonFee(
                    member_id=member.id,
                    season_label=period.season_label,
                    membership_fee=to_decimal(member.membership_fee),
                )
            )
    db.session.commit()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/assets/<path:filename>")
def asset_file(filename):
    return send_from_directory(BASE_DIR, filename)


@app.route("/api/config")
def config():
    return jsonify({"expense_categories": EXPENSE_CATEGORIES})


@app.route("/api/member-statuses")
@token_required({"board", "admin"})
def member_statuses():
    statuses = MemberStatus.query.order_by(MemberStatus.id.asc()).all()
    return jsonify({"statuses": [s.name for s in statuses]})


@app.route("/api/auth/init", methods=["POST"])
def auth_init():
    data = request.get_json() or {}
    phone = (data.get("phone") or "").strip()

    if not validate_phone(phone):
        return jsonify({"error": "Telefona Nr. jābūt ar 8 cipariem"}), 400

    user = Member.query.filter_by(phone=phone).first()
    if not user:
        return jsonify({"error": "Jūs neesat biedrs"}), 404
    if normalize_role(user.role) == "admin":
        return jsonify({"error": "Jūs neesat biedrs"}), 404
    if normalize_role(user.role) == "member":
        return jsonify({"error": "Jums nav tiesību ienākt."}), 403

    return jsonify({"needs_pin_setup": user.pin_hash is None})


@app.route("/api/auth/setup-pin", methods=["POST"])
def setup_pin():
    data = request.get_json() or {}
    phone = (data.get("phone") or "").strip()
    pin = (data.get("pin") or "").strip()
    pin_confirm = (data.get("pin_confirm") or "").strip()

    if not validate_phone(phone):
        return jsonify({"error": "Telefona Nr. jābūt ar 8 cipariem"}), 400
    if not validate_pin(pin):
        return jsonify({"error": "PIN kodam jābūt ar 4 cipariem"}), 400
    if pin != pin_confirm:
        return jsonify({"error": "PIN kodi nesakrīt"}), 400

    user = Member.query.filter_by(phone=phone).first()
    if not user:
        return jsonify({"error": "Jūs neesat biedrs"}), 404
    if normalize_role(user.role) == "member":
        return jsonify({"error": "Jums nav tiesību ienākt."}), 403
    if user.pin_hash:
        return jsonify({"error": "PIN kods jau ir uzstādīts"}), 409

    user.pin_hash = generate_password_hash(pin)
    db.session.commit()
    return jsonify({"message": "PIN kods veiksmīgi saglabāts"})


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    phone = (data.get("phone") or "").strip()
    pin = (data.get("pin") or "").strip()

    if not validate_phone(phone):
        return jsonify({"error": "Telefona Nr. jābūt ar 8 cipariem"}), 400

    user = Member.query.filter_by(phone=phone).first()
    if not user:
        return jsonify({"error": "Jūs neesat biedrs"}), 404
    if normalize_role(user.role) == "member":
        return jsonify({"error": "Jums nav tiesību ienākt."}), 403
    if not user.pin_hash:
        return jsonify({"error": "Lūdzu vispirms uzstādiet PIN kodu"}), 409
    if not validate_pin(pin) or not check_password_hash(user.pin_hash, pin):
        return jsonify({"error": "Nepareizs PIN kods"}), 401

    token = str(uuid.uuid4())
    TOKENS[token] = user.id

    return jsonify(
        {
            "token": token,
            "user": member_to_dict(user),
        }
    )


@app.route("/api/auth/logout", methods=["POST"])
@token_required()
def logout():
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "").strip()
    TOKENS.pop(token, None)
    return jsonify({"message": "Izrakstīšanās izdevusies"})


@app.route("/api/dashboard")
@token_required()
def dashboard():
    period = get_period_for_request()
    totals = period_totals(period)

    return jsonify(
        {
            "period": {
                "season_label": period.season_label,
                "start_date": period.start_date.isoformat(),
                "end_date": period.end_date.isoformat(),
                "carry_over": float(period.carry_over),
            },
            "totals": totals,
        }
    )


@app.route("/api/members", methods=["GET"])
@token_required({"board", "admin", "auditor", "cashier", "member"})
def list_members():
    if ensure_unique_member_numbers():
        db.session.commit()

    period = get_period_for_request()
    requester_role = normalize_role(request.current_user.role)
    members = Member.query.order_by(Member.list_no).all()
    if requester_role != "admin":
        members = [m for m in members if normalize_role(m.role) != "admin"]

    paid_rows = (
        db.session.query(
            Income.member_id,
            db.func.coalesce(db.func.sum(Income.amount), 0),
        )
        .filter(
            Income.income_type.in_(["member_fee", "biedra nauda"]),
            Income.member_id.isnot(None),
            Income.entry_date >= period.start_date,
            Income.entry_date <= period.end_date,
        )
        .group_by(Income.member_id)
        .all()
    )
    paid_map = {member_id: total for member_id, total in paid_rows}
    fee_rows = MemberSeasonFee.query.filter_by(season_label=period.season_label).all()
    fee_map = {row.member_id: row.membership_fee for row in fee_rows}

    return jsonify(
        [
            member_to_dict(
                m,
                paid_this_period_override=paid_map.get(m.id, 0),
                membership_fee_override=fee_map.get(m.id, m.membership_fee),
            )
            for m in members
        ]
    )


@app.route("/api/members", methods=["POST"])
@token_required({"board", "admin"})
def create_member():
    data = request.get_json() or {}

    phone = str(data.get("phone", "")).strip()
    role = normalize_role(data.get("role", "member"))
    season_label = str(data.get("season_label", current_period().season_label)).strip()
    membership_fee = to_decimal(data.get("membership_fee", 0) or 0)

    if not validate_phone(phone):
        return jsonify({"error": "Telefona Nr. jābūt ar 8 cipariem"}), 400
    if Member.query.filter_by(phone=phone).first():
        return jsonify({"error": "Lietotājs ar šo telefona numuru jau eksistē"}), 409
    if role not in ROLES:
        return jsonify({"error": "Nederīga loma"}), 400
    if normalize_role(request.current_user.role) != "admin" and role == "admin":
        return jsonify({"error": "Tikai admin drikst izveidot admin kontu"}), 403

    member = Member(
        list_no=next_member_list_no(),
        first_name=str(data.get("first_name", "")).strip() or "Vards",
        last_name=str(data.get("last_name", "")).strip() or "Uzvards",
        phone=phone,
        status=str(data.get("status", "active")).strip(),
        membership_fee=membership_fee,
        paid_this_period=0,
        joining_fee_paid=to_bool(data.get("joining_fee_paid", False)),
        role=role,
    )

    db.session.add(member)
    db.session.flush()
    set_membership_fee_for_season(member, season_label, membership_fee)
    db.session.commit()

    return jsonify(member_to_dict(member)), 201


@app.route("/api/members/<int:member_id>", methods=["PUT"])
@token_required({"board", "admin"})
def update_member(member_id):
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"error": "Biedrs nav atrasts"}), 404

    requester_role = normalize_role(request.current_user.role)
    if requester_role != "admin" and is_admin_member(member):
        return jsonify({"error": "Admin konts ir redzams un labojams tikai admin"}), 403

    data = request.get_json() or {}
    phone = str(data.get("phone", member.phone)).strip()
    season_label = str(data.get("season_label") or request.args.get("season_label") or current_period().season_label).strip()

    if not validate_phone(phone):
        return jsonify({"error": "Telefona Nr. jābūt ar 8 cipariem"}), 400

    duplicate = Member.query.filter(Member.phone == phone, Member.id != member_id).first()
    if duplicate:
        return jsonify({"error": "Telefona numurs jau aizņemts"}), 409

    member.first_name = str(data.get("first_name", member.first_name)).strip()
    member.last_name = str(data.get("last_name", member.last_name)).strip()
    member.phone = phone
    member.status = str(data.get("status", member.status)).strip()
    membership_fee = to_decimal(data.get("membership_fee", get_effective_membership_fee(member, season_label)))
    set_membership_fee_for_season(member, season_label, membership_fee)
    if current_period().season_label == season_label:
        member.membership_fee = membership_fee
    member.joining_fee_paid = to_bool(data.get("joining_fee_paid", member.joining_fee_paid))

    if requester_role == "admin":
        role = normalize_role(data.get("role", member.role))
        if role not in ROLES:
            return jsonify({"error": "Nederīga loma"}), 400
        member.role = role

    db.session.commit()
    return jsonify(member_to_dict(member, membership_fee_override=membership_fee))


@app.route("/api/members/<int:member_id>", methods=["DELETE"])
@token_required({"board", "admin"})
def delete_member(member_id):
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"error": "Biedrs nav atrasts"}), 404

    if normalize_role(request.current_user.role) != "admin" and is_admin_member(member):
        return jsonify({"error": "Admin kontu var dzest tikai admin"}), 403

    db.session.delete(member)
    db.session.commit()

    resequence_members()
    db.session.commit()

    return jsonify({"message": "Biedrs izdzēsts"})


@app.route("/api/members/<int:member_id>/pin", methods=["DELETE"])
@token_required({"admin"})
def clear_member_pin(member_id):
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"error": "Biedrs nav atrasts"}), 404

    if member.pin_hash is None:
        return jsonify({"error": "PIN kods jau ir dzests"}), 409

    member.pin_hash = None
    db.session.commit()

    return jsonify({"message": "PIN kods izdzests"})


@app.route("/api/members/<int:member_id>/payment", methods=["POST"])
@token_required({"cashier", "admin"})
def record_member_payment(member_id):
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"error": "Biedrs nav atrasts"}), 404

    if normalize_role(request.current_user.role) != "admin" and is_admin_member(member):
        return jsonify({"error": "Admin konts nav pieejams"}), 403

    data = request.get_json() or {}
    amount = to_decimal(data.get("amount", 0) or 0)
    entry_date = datetime.strptime(data.get("entry_date", date.today().isoformat()), "%Y-%m-%d").date()

    if amount <= 0:
        return jsonify({"error": "Summai jābūt lielākai par 0"}), 400

    income = Income(
        income_type="member_fee",
        member_id=member.id,
        amount=amount,
        entry_date=entry_date,
        description="Biedra naudas iemaksa",
    )
    member.paid_this_period = to_decimal(member.paid_this_period) + amount

    db.session.add(income)
    db.session.commit()

    active_season_label = current_period().season_label
    current_membership_fee = get_effective_membership_fee(member, active_season_label)
    progress = 0.0
    if current_membership_fee > 0:
        progress = float((to_decimal(member.paid_this_period) / current_membership_fee) * 100)

    return jsonify({"message": "Iemaksa pievienota", "progress_percent": round(progress, 2)})


@app.route("/api/incomes", methods=["POST"])
@token_required({"cashier", "admin"})
def add_other_income():
    data = request.get_json() or {}
    amount = to_decimal(data.get("amount", 0) or 0)
    entry_date = datetime.strptime(data.get("entry_date", date.today().isoformat()), "%Y-%m-%d").date()
    description = str(data.get("description", "")).strip()

    if amount <= 0:
        return jsonify({"error": "Summai jābūt lielākai par 0"}), 400

    income = Income(
        income_type="neplānots ienākums",
        amount=amount,
        entry_date=entry_date,
        description=description,
    )
    db.session.add(income)
    db.session.commit()
    return jsonify({"message": "Ieņēmumi pievienoti"}), 201


@app.route("/api/expenses", methods=["POST"])
@token_required({"cashier", "admin", "member"})
def add_expense():
    category = (request.form.get("category") or "").strip()
    amount_raw = request.form.get("amount", "0")
    entry_date_raw = request.form.get("entry_date", date.today().isoformat())
    description = (request.form.get("description") or "").strip()

    if category not in EXPENSE_CATEGORIES:
        return jsonify({"error": "Nederīga izdevumu kategorija"}), 400

    amount = to_decimal(amount_raw)
    if amount <= 0:
        return jsonify({"error": "Summai jābūt lielākai par 0"}), 400

    entry_date = datetime.strptime(entry_date_raw, "%Y-%m-%d").date()

    attachment_name = None
    file = request.files.get("attachment")
    if file and file.filename:
        safe_name = secure_filename(file.filename)
        unique_name = f"{uuid.uuid4()}_{safe_name}"
        file.save(UPLOAD_DIR / unique_name)
        attachment_name = unique_name

    expense = Expense(
        category=category,
        amount=amount,
        entry_date=entry_date,
        description=description,
        attachment=attachment_name,
        created_by_member_id=request.current_user.id,
    )

    db.session.add(expense)
    db.session.commit()

    return jsonify({"message": "Izdevumi pievienoti"}), 201


@app.route("/api/history")
@token_required({"cashier", "board", "admin", "auditor"})
def history():
    period = get_period_for_request()
    requester_role = normalize_role(request.current_user.role)

    incomes = Income.query.filter(
        Income.entry_date >= period.start_date,
        Income.entry_date <= period.end_date,
    ).order_by(Income.entry_date.desc()).all()

    expenses = Expense.query.filter(
        Expense.entry_date >= period.start_date,
        Expense.entry_date <= period.end_date,
    ).order_by(Expense.entry_date.desc()).all()

    income_rows = []
    for row in incomes:
        member = db.session.get(Member, row.member_id) if row.member_id else None
        member_name = ""
        if member and (requester_role == "admin" or normalize_role(member.role) != "admin"):
            member_name = f"{member.first_name} {member.last_name}"
        income_rows.append(
            {
                "id": row.id,
                "type": row.income_type,
                "member_name": member_name,
                "amount": float(row.amount),
                "entry_date": row.entry_date.isoformat(),
                "description": row.description or "",
            }
        )

    expense_rows = []
    for row in expenses:
        expense_rows.append(
            {
                "id": row.id,
                "category": row.category,
                "amount": float(row.amount),
                "entry_date": row.entry_date.isoformat(),
                "description": row.description or "",
                "attachment": row.attachment,
            }
        )

    return jsonify({"incomes": income_rows, "expenses": expense_rows})


@app.route("/api/attachments/<path:filename>")
@token_required({"cashier", "board", "admin", "auditor", "member"})
def get_attachment(filename):
    return send_from_directory(UPLOAD_DIR, filename, as_attachment=False)


@app.route("/api/periods/available")
@token_required()
def available_periods():
    periods = Period.query.order_by(Period.start_date.desc()).all()
    if not periods:
        periods = [current_period()]

    unique = []
    seen = set()
    for period in periods:
        if period.season_label in seen:
            continue
        seen.add(period.season_label)
        unique.append(
            {
                "season_label": period.season_label,
                "start_date": period.start_date.isoformat(),
                "end_date": period.end_date.isoformat(),
                "active": bool(period.active),
            }
        )

    return jsonify({"periods": unique})


@app.route("/api/period", methods=["POST"])
@token_required({"board", "admin"})
def update_period():
    data = request.get_json() or {}
    season_label = (data.get("season_label") or "").strip()
    default_membership_fee = to_decimal(data.get("default_membership_fee", 0) or 0)
    carry_over = to_decimal(data.get("carry_over", 0) or 0)
    current_user_role = normalize_role(request.current_user.role)
    active_period = Period.query.filter_by(active=True).first()
    is_new_season = active_period is None or active_period.season_label != season_label

    period_lock = get_or_create_period_lock(season_label)
    if current_user_role == "board" and period_lock.membership_fee_locked:
        return jsonify({"error": "Saja perioda biedra maksu pec pirmas saglabasanas var mainit tikai admin"}), 403
    if current_user_role == "board" and period_lock.carry_over_locked:
        return jsonify({"error": "Saja perioda atlikumu pec pirmas saglabasanas var mainit tikai admin"}), 403

    if len(season_label) != 9 or "/" not in season_label:
        return jsonify({"error": "Sezonas formatam jabut yyyy/yyyy"}), 400

    start_year = int(season_label[:4])
    end_year = int(season_label[5:])
    if end_year != start_year + 1:
        return jsonify({"error": "Sezonas gadi neatbilst formatam yyyy/yyyy"}), 400

    if is_new_season:
        Period.query.update({"active": False})
        period = Period(
            season_label=season_label,
            start_date=date(start_year, 4, 1),
            end_date=date(end_year, 3, 31),
            carry_over=0,
            active=True,
        )
        db.session.add(period)
    else:
        period = active_period
        period.start_date = date(start_year, 4, 1)
        period.end_date = date(end_year, 3, 31)
        period.active = True
    period.carry_over = carry_over

    members = Member.query.all()

    if default_membership_fee > 0:
        for member in members:
            fee_value = calculate_membership_fee_for_period(member, default_membership_fee)
            set_membership_fee_for_season(member, season_label, fee_value)
            member.membership_fee = fee_value
            member.paid_this_period = 0

        if current_user_role == "board":
            period_lock.membership_fee_locked = True

    if current_user_role == "board":
        period_lock.carry_over_locked = True

    # Iestāšanās maksas atzīme ir aktīva tikai vienu sezonu.
    if is_new_season:
        for member in members:
            member.joining_fee_paid = False

    db.session.commit()
    return jsonify({"message": "Parskata periods atjaunots"})


@app.route("/api/period/carryover", methods=["POST"])
@token_required({"board", "admin"})
def set_carryover():
    data = request.get_json() or {}
    carry_over = to_decimal(data.get("carry_over", 0) or 0)
    current_user_role = normalize_role(request.current_user.role)

    period = current_period()
    period_lock = get_or_create_period_lock(period.season_label)
    if current_user_role == "board" and period_lock.carry_over_locked:
        return jsonify({"error": "Saja perioda atlikumu pec pirmas saglabasanas var mainit tikai admin"}), 403

    period.carry_over = carry_over

    if current_user_role == "board":
        period_lock.carry_over_locked = True

    db.session.commit()

    return jsonify({"message": "Atlikums no iepriekseja perioda saglabats"})


@app.route("/api/export")
@token_required({"cashier", "board", "admin", "auditor"})
def export_balance():
    period = get_period_for_request()
    requester_role = normalize_role(request.current_user.role)
    totals = period_totals(period)

    incomes = Income.query.filter(
        Income.entry_date >= period.start_date,
        Income.entry_date <= period.end_date,
    ).order_by(Income.entry_date.asc()).all()

    expenses = Expense.query.filter(
        Expense.entry_date >= period.start_date,
        Expense.entry_date <= period.end_date,
    ).order_by(Expense.entry_date.asc()).all()

    wb = Workbook()
    ws_income = wb.active
    ws_income.title = "Ienemumi"

    ws_income.append(["Tips", "Biedrs", "Summa EUR", "Datums", "Apraksts"])
    ws_income.append(["Atlikums no iepriekseja perioda", "", float(period.carry_over), "", ""])

    for row in incomes:
        member = db.session.get(Member, row.member_id) if row.member_id else None
        member_name = ""
        if member and (requester_role == "admin" or normalize_role(member.role) != "admin"):
            member_name = f"{member.first_name} {member.last_name}"
        ws_income.append(
            [
                row.income_type,
                member_name,
                float(row.amount),
                row.entry_date.isoformat(),
                row.description or "",
            ]
        )

    ws_income.append(["Ienemumi kopā EUR", "", totals["income_total"], "", ""])

    ws_expense = wb.create_sheet("Izdevumi")
    ws_expense.append(["Kategorija", "Summa EUR", "Datums", "Apraksts", "Pievienotais fails"])

    for row in expenses:
        ws_expense.append(
            [
                row.category,
                float(row.amount),
                row.entry_date.isoformat(),
                row.description or "",
                row.attachment or "",
            ]
        )

    ws_expense.append(["Izdevumi kopā EUR", totals["expense_total"], "", "", ""])

    ws_summary = wb.create_sheet("Kopsavilkums")
    ws_summary.append(["Pārskata periods", period.season_label])
    ws_summary.append(["Ieņēmumi EUR", totals["income_total"]])
    ws_summary.append(["Izdevumi EUR", totals["expense_total"]])
    if totals["difference"] >= 0:
        ws_summary.append(["Atlikums EUR", totals["balance"]])
    else:
        ws_summary.append(["Deficīts EUR", totals["deficit"]])

    filename = f"bilance_{period.season_label.replace('/', '-')}.xlsx"
    file_path = BASE_DIR / filename
    wb.save(file_path)

    return send_from_directory(BASE_DIR, filename, as_attachment=True)


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        ensure_seed_data()

    app.run(
        host=os.getenv("FLASK_RUN_HOST", "0.0.0.0"),
        port=int(os.getenv("FLASK_RUN_PORT", "5000")),
        debug=os.getenv("FLASK_DEBUG", "1") == "1",
    )
